# 听写单词，短语，常见句子
import glob
import os
import random
import openpyxl

from gtts import gTTS
from playsound import playsound
# from pydub import AudioSegment


# 1. 听写单词

# 音频播放速度控制函数
def audio_speed_controller(sound, speed=1.0):
    # modified_sound = sound._spawn(sound.raw_data, overrides={
    #     "frame_rate": int(sound.frame_rate * speed)
    # })
    # return modified_sound.set_frame_rate(sound.frame_rate)
    return sound


# 删除音频缓存函数
def delete_temp_audio(delete_directory=os.path.join(os.getcwd(), 'temp')):
    for file_path in glob.glob(os.path.join(delete_directory, '*.mp3')):
        if os.path.isfile(file_path):
            os.remove(file_path)
            # print(f"Deleted: {file_path}")


# 生成音频函数
def generate_mp3file(chosen_text="おはよう，日本！", sound_speed=1.0, mp3file_path=""):
    tts_mp3file = gTTS(text=chosen_text, lang='ja', slow=sound_speed == 0.9)
    tts_mp3file.save(mp3file_path)

    # If Sound Speed Need Adjust
    if sound_speed != 1.0 and sound_speed != 0.9:
        sound = AudioSegment.from_file(mp3file_path)
        adjusted_speed_sound = audio_speed_controller(sound, speed=sound_speed)
        adjusted_speed_sound.export(mp3file_path, format="mp3")


# 播放单个音频函数
def play_single_mp3(os_cwd="", mp3file_path="temp_audio.mp3"):
    sound_playing = os.path.join(os_cwd, mp3file_path)
    # print(f"Playing: {sound_playing}")
    playsound(sound_playing)


# 选择听力内容函数
def choose_content(db_data):
    chosen_text = random.choice(db_data)
    return chosen_text


# Function1: 加载日付、時間到data_list
def load_kanji_and_hiragana(file_path, sheet_name):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 提取 HIRAGANA 和 KANJI 列的数据 (假设 HIRAGANA 在第 1 列，KANJI 在第 2 列)
    data_list = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        hiragana = row[0].value  # 第 1 列是 HIRAGANA
        kanji = row[1].value  # 第 2 列是 KANJI
        if hiragana and kanji:  # 确保两列都不为空
            data_list.append({"hiragana": hiragana, "kanji": kanji})

    return data_list


# Function2: 加载名詞到data_list
def load_meishi(file_path, sheet_name):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 提取 MONO, NAME, KANA, KANJI, ENGLISH列的数据
    data_list = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1):
        mono = row[0].value  # 第 1 列是 MONO
        name = row[1].value  # 第 2 列是 NAME，发音发这个
        kana = row[2].value  # 第 3 列是 KANA，发音发这个
        kanji = row[3].value  # 第 4 列是 KANJI，发音发这个
        english = row[4].value  # 第 5 列是 ENGLISH，发音发这个
        if mono or name or kanji or kana or english:  # 确保都不为空
            data_list.append({
                "mono": mono,
                "name": name,
                "kana": kana,
                "kanji": kanji,
                "english": english,
            })

    return data_list
